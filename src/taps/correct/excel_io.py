"""
Excel I/O 모듈 (검수 워크플로우)

export_issues_to_xlsx: issues.jsonl → review.xlsx
import_xlsx_to_resolutions: review.xlsx → resolutions.jsonl + text_avail_final.jsonl

CLI:
    # Export
    python -m taps.correct.excel_io export \\
        --issues_jsonl issues.jsonl \\
        --output_xlsx review.xlsx

    # Import
    python -m taps.correct.excel_io import \\
        --input_xlsx review.xlsx \\
        --output_resolutions resolutions.jsonl \\
        --output_text_avail text_avail_final.jsonl
"""

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Any

from .models import Issue, Candidate


# =============================================================================
# Export: Issues -> Excel
# =============================================================================

def export_issues_to_xlsx(
    issues_jsonl: str,
    output_xlsx: str,
    max_candidates_display: int = 5,
) -> int:
    """
    Issues를 Excel로 내보내기

    컬럼 (implementation_contract_v1.md 준수):
    - utt_id, speaker_id, sentence_id
    - bucket, tag
    - span_start, span_end, raw_span
    - context_marked
    - candidates (stringify)
    - recommended
    - user_fix (prefill = recommended)
    - avg_logprob, compression_ratio

    Args:
        issues_jsonl: 입력 issues JSONL 파일
        output_xlsx: 출력 Excel 파일
        max_candidates_display: 표시할 최대 후보 수

    Returns:
        int: 내보낸 이슈 수
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    # Issues 로드
    issues: List[Issue] = []
    with open(issues_jsonl, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                issues.append(Issue.from_dict(json.loads(line)))

    if not issues:
        print(f"경고: {issues_jsonl}에 이슈가 없습니다.")
        return 0

    # Workbook 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Issues"

    # 헤더
    headers = [
        "utt_id",
        "speaker_id",
        "sentence_id",
        "bucket",
        "tag",
        "span_start",
        "span_end",
        "raw_span",
        "context_marked",
        "candidates",
        "recommended",
        "user_fix",
        "avg_logprob",
        "compression_ratio",
    ]

    # 스타일 정의
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 버킷별 색상
    bucket_colors = {
        "RED": "FF6B6B",
        "ORANGE": "FFA94D",
        "YELLOW": "FFD93D",
        "GREEN": "6BCB77",
    }

    # 헤더 작성
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 데이터 행 작성
    for row_idx, issue in enumerate(issues, 2):
        # candidates를 문자열로 (score 포함)
        cands_str = " | ".join(
            f"{c.text} ({c.score:.3f})"
            for c in issue.candidates[:max_candidates_display]
        )

        row_data = [
            issue.utt_id,
            issue.speaker_id,
            issue.sentence_id,
            issue.bucket,
            issue.tag,
            issue.span_start,
            issue.span_end,
            issue.raw_span,
            issue.context_marked,
            cands_str,
            issue.recommended,
            issue.user_fix,
            issue.meta.get("avg_logprob", ""),
            issue.meta.get("compression_ratio", ""),
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 9))

            # bucket 열에 색상 적용
            if col == 4 and issue.bucket in bucket_colors:
                cell.fill = PatternFill(
                    start_color=bucket_colors[issue.bucket],
                    end_color=bucket_colors[issue.bucket],
                    fill_type="solid",
                )

    # 열 너비 조정
    column_widths = {
        "A": 22,  # utt_id
        "B": 12,  # speaker_id
        "C": 12,  # sentence_id
        "D": 10,  # bucket
        "E": 8,   # tag
        "F": 10,  # span_start
        "G": 10,  # span_end
        "H": 20,  # raw_span
        "I": 50,  # context_marked
        "J": 45,  # candidates
        "K": 20,  # recommended
        "L": 20,  # user_fix
        "M": 12,  # avg_logprob
        "N": 15,  # compression_ratio
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 첫 행 고정
    ws.freeze_panes = "A2"

    # 저장
    Path(output_xlsx).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    print(f"Excel 내보내기 완료: {output_xlsx} ({len(issues)}개 이슈)")

    return len(issues)


# =============================================================================
# Import: Excel -> Resolutions
# =============================================================================

def import_xlsx_to_resolutions(
    input_xlsx: str,
    output_resolutions_jsonl: str,
    output_text_avail_final_jsonl: str,
    original_asr_jsonl: Optional[str] = None,
) -> Tuple[int, int]:
    """
    검수 완료된 Excel을 역변환

    Args:
        input_xlsx: 검수 완료된 Excel 파일
        output_resolutions_jsonl: Resolution 출력 JSONL
        output_text_avail_final_jsonl: 최종 text_avail 출력 JSONL
        original_asr_jsonl: 원본 ASR JSONL (text_raw 복원용)

    Returns:
        (resolutions_count, text_avail_count)

    Import 규칙:
    - user_fix가 비어있으면 recommended로 간주
    - 문장 단위로 병합하여 text_avail 생성
    """
    import openpyxl

    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active

    # 헤더 읽기
    headers = [cell.value for cell in ws[1]]

    # 원본 ASR 로드 (있으면)
    original_texts: Dict[str, str] = {}
    if original_asr_jsonl and Path(original_asr_jsonl).exists():
        with open(original_asr_jsonl, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    rec = json.loads(line)
                    utt_id = rec.get("utt_id", f"{rec.get('speaker_id', '')}_{rec.get('sentence_id', '')}")
                    text = rec.get("text", rec.get("text_raw", ""))
                    original_texts[utt_id] = text

    # Excel 데이터 읽기
    resolutions: List[Dict[str, Any]] = []
    issues_by_utt: Dict[str, List[Dict[str, Any]]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # 빈 행 스킵
            continue

        row_dict = dict(zip(headers, row))

        utt_id = str(row_dict.get("utt_id", ""))
        user_fix = row_dict.get("user_fix", "")
        recommended = row_dict.get("recommended", "")

        # user_fix가 비어있으면 recommended 사용
        if user_fix is None or str(user_fix).strip() == "":
            final_text = str(recommended) if recommended else ""
        else:
            final_text = str(user_fix)

        resolution = {
            "utt_id": utt_id,
            "speaker_id": str(row_dict.get("speaker_id", "")),
            "sentence_id": str(row_dict.get("sentence_id", "")),
            "span_start": int(row_dict.get("span_start", 0) or 0),
            "span_end": int(row_dict.get("span_end", 0) or 0),
            "raw_span": str(row_dict.get("raw_span", "")),
            "final_text": final_text,
            "was_modified": (
                user_fix is not None and
                str(user_fix).strip() != "" and
                str(user_fix) != str(recommended)
            ),
            "resolved_at": datetime.now().isoformat(),
        }
        resolutions.append(resolution)

        if utt_id not in issues_by_utt:
            issues_by_utt[utt_id] = []
        issues_by_utt[utt_id].append(resolution)

    # Resolutions 저장
    Path(output_resolutions_jsonl).parent.mkdir(parents=True, exist_ok=True)
    with open(output_resolutions_jsonl, "w", encoding="utf-8") as f:
        for res in resolutions:
            f.write(json.dumps(res, ensure_ascii=False) + "\n")

    # text_avail_final 생성 (스팬 적용)
    text_avail_finals: List[Dict[str, Any]] = []

    for utt_id, issue_resolutions in issues_by_utt.items():
        if not issue_resolutions:
            continue

        # 원본 텍스트 가져오기
        if utt_id in original_texts:
            text_base = original_texts[utt_id]
        else:
            # context_full 또는 raw_span으로 추론
            # 가장 긴 raw_span이 전체 문장인 경우를 처리
            max_span = max(issue_resolutions, key=lambda r: r["span_end"] - r["span_start"])
            if max_span["span_start"] == 0:
                text_base = max_span["raw_span"]
            else:
                # 복원 불가 - 경고 출력
                print(f"경고: {utt_id}의 원본 텍스트를 복원할 수 없습니다.")
                continue

        # 스팬 적용 (역순으로)
        sorted_res = sorted(
            issue_resolutions,
            key=lambda r: r["span_start"],
            reverse=True,
        )

        text_final = text_base
        for res in sorted_res:
            start = res["span_start"]
            end = res["span_end"]
            text_final = text_final[:start] + res["final_text"] + text_final[end:]

        final_rec = {
            "utt_id": utt_id,
            "speaker_id": issue_resolutions[0]["speaker_id"],
            "sentence_id": issue_resolutions[0]["sentence_id"],
            "text_raw": text_base,
            "text_avail": text_final,
            "resolved_from": "human_review",
            "resolution_count": len(issue_resolutions),
        }
        text_avail_finals.append(final_rec)

    # 저장
    Path(output_text_avail_final_jsonl).parent.mkdir(parents=True, exist_ok=True)
    with open(output_text_avail_final_jsonl, "w", encoding="utf-8") as f:
        for rec in text_avail_finals:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    print(f"Import 완료:")
    print(f"  Resolutions: {len(resolutions)}개 -> {output_resolutions_jsonl}")
    print(f"  text_avail_final: {len(text_avail_finals)}개 -> {output_text_avail_final_jsonl}")

    return len(resolutions), len(text_avail_finals)


# =============================================================================
# apply_resolutions (implementation_contract_v1.md 시그니처)
# =============================================================================

def apply_resolutions(
    text_raw: str,
    issues: List[Issue],
    resolved_user_fixes: Dict[Tuple[int, int], str],
) -> str:
    """
    스팬에 사용자 수정 적용

    Args:
        text_raw: 원본 텍스트
        issues: 이슈 리스트
        resolved_user_fixes: {(start, end): user_fix} 딕셔너리

    Returns:
        수정된 텍스트

    겹침 처리: 더 긴 스팬 우선, 같으면 먼저 시작하는 스팬 우선
    """
    if not resolved_user_fixes:
        return text_raw

    # (start, end, new_text) 리스트로 변환
    fixes = [
        (start, end, new_text)
        for (start, end), new_text in resolved_user_fixes.items()
    ]

    # 역순 정렬 (뒤에서부터 적용)
    sorted_fixes = sorted(fixes, key=lambda f: (f[0], -(f[1] - f[0])), reverse=True)

    result = text_raw
    applied_ranges: List[Tuple[int, int]] = []

    for start, end, new_text in sorted_fixes:
        # 겹침 체크
        overlaps = False
        for app_start, app_end in applied_ranges:
            if not (end <= app_start or start >= app_end):
                overlaps = True
                break

        if not overlaps:
            result = result[:start] + new_text + result[end:]
            applied_ranges.append((start, end))

    return result


# =============================================================================
# CLI
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Excel I/O for correct_model (검수 워크플로우)"
    )
    subparsers = parser.add_subparsers(dest="command", help="명령어")

    # Export 명령
    export_parser = subparsers.add_parser("export", help="Issues -> Excel")
    export_parser.add_argument(
        "--issues_jsonl",
        required=True,
        help="입력 issues JSONL 파일",
    )
    export_parser.add_argument(
        "--output_xlsx",
        required=True,
        help="출력 Excel 파일",
    )
    export_parser.add_argument(
        "--max_candidates",
        type=int,
        default=5,
        help="표시할 최대 후보 수 (기본: 5)",
    )

    # Import 명령
    import_parser = subparsers.add_parser("import", help="Excel -> Resolutions")
    import_parser.add_argument(
        "--input_xlsx",
        required=True,
        help="검수 완료된 Excel 파일",
    )
    import_parser.add_argument(
        "--output_resolutions",
        required=True,
        help="Resolution 출력 JSONL",
    )
    import_parser.add_argument(
        "--output_text_avail",
        required=True,
        help="최종 text_avail 출력 JSONL",
    )
    import_parser.add_argument(
        "--original_asr",
        default=None,
        help="원본 ASR JSONL (text_raw 복원용)",
    )

    args = parser.parse_args()

    if args.command == "export":
        export_issues_to_xlsx(
            args.issues_jsonl,
            args.output_xlsx,
            max_candidates_display=args.max_candidates,
        )
    elif args.command == "import":
        import_xlsx_to_resolutions(
            args.input_xlsx,
            args.output_resolutions,
            args.output_text_avail,
            original_asr_jsonl=args.original_asr,
        )
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
